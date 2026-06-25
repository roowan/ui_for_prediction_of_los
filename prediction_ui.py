"""
EHR Workup Time Predictor — Desktop GUI (XGBoost)
==================================================
Dataset: ehr_age7_ar1.xlsx
Run:     python predict_app_age7_ar1.py
Requires: workup_model_age7_ar1.pkl in the same folder

Changes in this version
-----------------------
• Session / Arrival-slot validation:
    Forenoon  → 08:00-09:00 … 12:00-13:00  (start hour 8-12, before 1 PM)
    Lunch     → 13:00-14:00 … 14:00-15:00  (start hour 13-14, 1 PM – 3 PM)
    Afternoon → 15:00-16:00 … 16:00-17:00  (start hour 15-16, after 3 PM)
  Mismatch shows an inline error banner — prediction is blocked.

• New "Bulk Predict" tab:
    – Browse for an Excel (.xlsx) file
    – Required columns mapped & validated
    – Runs predictions for all rows (background thread)
    – Shows results in a scrollable table
    – Export results to Excel button
"""

import pickle, math, threading, os
import numpy as np
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font as tkfont

# ── Palette ────────────────────────────────────────────────────────
BG         = "#080C14"
PANEL      = "#0E1420"
CARD       = "#131A28"
CARD2      = "#192030"
BORDER     = "#1E2D45"
BORDER2    = "#243452"
ACCENT     = "#3D8EF5"
ACCENT2    = "#6C4FED"
ACCENT_GLO = "#1A3A6E"
TEXT       = "#E8EDF8"
TEXT_DIM   = "#6A7898"
TEXT_MUTED = "#374060"
SUCCESS    = "#2ED87A"
SUCCESS_DIM= "#1A4A35"
WARN       = "#F5A623"
WARN_DIM   = "#4A3310"
DANGER     = "#F5504A"
DANGER_DIM = "#4A1A18"
WHITE      = "#FFFFFF"
ERROR_BG   = "#2D1A1A"
ERROR_BDR  = "#7A2020"

# ── Session → valid arrival start-hours ───────────────────────────
#   Forenoon  : before 1 PM  → hours 8, 9, 10, 11, 12
#   Lunch     : 1 PM – 3 PM  → hours 13, 14
#   Afternoon : after 3 PM   → hours 15, 16
SESSION_VALID_HOURS = {
    "Forenoon":  {8, 9, 10, 11, 12},
    "Lunch":     {13, 14},
    "Afternoon": {15, 16},
}
87
SESSION_HINT = {
    "Forenoon":  "Forenoon covers 08:00 – 13:00 (before 1 PM).",
    "Lunch":     "Lunch hour covers 13:00 – 15:00 (1 PM – 3 PM).",
    "Afternoon": "Afternoon covers 15:00 – 17:00 (after 3 PM).",
}

# Derive session automatically from arrival slot (single-patient tab)
SESSION_BADGE_COLOR = {
    "Forenoon":  "#1A4A35",
    "Lunch":     "#4A3A10",
    "Afternoon": "#0E2E5A",
}

def arrival_to_session(arrival_slot: str) -> str:
    hour = int(ARRIVAL_MAP.get(arrival_slot, 9))
    if hour in SESSION_VALID_HOURS["Forenoon"]:  return "Forenoon"
    if hour in SESSION_VALID_HOURS["Lunch"]:     return "Lunch"
    return "Afternoon"

# ── Arrival map: 1-hour slots → numeric start hour ────────────────
ARRIVAL_MAP = {
    "08:00 - 09:00": 8.0,
    "09:00 - 10:00": 9.0,
    "10:00 - 11:00": 10.0,
    "11:00 - 12:00": 11.0,
    "12:00 - 13:00": 12.0,
    "13:00 - 14:00": 13.0,
    "14:00 - 15:00": 14.0,
    "15:00 - 16:00": 15.0,
    "16:00 - 17:00": 16.0,
}

# Reverse: start hour → slot label (for bulk validation feedback)
HOUR_TO_SLOT = {int(v): k for k, v in ARRIVAL_MAP.items()}

# ── Age category: display label → exact data string ───────────────
AGE_LABEL_MAP = {
    "0-12 (Paediatric)":   "0-12",
    "13-19 (Teenage)":     "13-19",
    "20-30 (Young Adult)": "20-30",
    "31-45 (Adult)":       "31-45",
    "46-60 (Middle Age)":  "46-60",
    "61-75 (Senior)":      "61-75",
    "76-90 (Elderly)":     "76-90",
}

# ── Age: numeric → category bin ───────────────────────────────────
def age_to_category(age) -> str:
    """Convert a numeric patient age to the model's age category string.
    Returns '31-45' as a safe default for unrecognisable values."""
    try:
        a = int(float(str(age).strip()))
    except (ValueError, TypeError):
        return "31-45"
    if   a <=  12: return "0-12"
    elif a <=  19: return "13-19"
    elif a <=  30: return "20-30"
    elif a <=  45: return "31-45"
    elif a <=  60: return "46-60"
    elif a <=  75: return "61-75"
    else:          return "76-90"

# ── Day full name → abbreviation (matches training data) ──────────
DAY_MAP = {
    "Monday": "Mon", "Tuesday": "Tue", "Wednesday": "Wed",
    "Thursday": "Thu", "Friday": "Fri", "Saturday": "Sat", "Sunday": "Sun",
}

# ── Bulk upload: expected column names (case-insensitive) ──────────
BULK_COL_MAP = {
    "age category":              "Age category",
    "age_category":              "Age category",
    "age":                       "Age",
    "patient age":               "Age",
    "patient_age":               "Age",
    "age (years)":               "Age",
    "age_years":                 "Age",
    "gender":                    "Gender",
    "patient visit category":    "Patient visit category",
    "visit_category":            "Patient visit category",
    "patient flow type":         "Patient Flow type",
    "flow_type":                 "Patient Flow type",
    "specialty":                 "Specialty",
    "consultant designation":    "Consultant Designation",
    "consultant_designation":    "Consultant Designation",
    "day":                       "Day",
    "session":                   "Session",
    "arrival hour":              "Arrival hour",
    "arrival_hour":              "Arrival hour",
    "arrival slot":              "Arrival hour",
    "arrival_slot":              "Arrival hour",
}

REQUIRED_BULK_COLS = [
    "Gender", "Patient visit category",
    "Patient Flow type", "Specialty",
    "Consultant Designation", "Day", "Session", "Arrival hour",
]
# At least one of these must be present for age
AGE_COLS_EITHER = ["Age category", "Age"]

# ── Model I/O ──────────────────────────────────────────────────────
def load_model(path="workup_model_age7_ar1.pkl"):
    with open(path, "rb") as f:
        return pickle.load(f)


def validate_session_arrival(session: str, arrival_slot: str) -> str | None:
    """Return error message string if mismatch, else None."""
    hour = int(ARRIVAL_MAP.get(arrival_slot, -1))
    if hour < 0:
        return f"Unknown arrival slot: '{arrival_slot}'"
    valid = SESSION_VALID_HOURS.get(session, set())
    if hour not in valid:
        valid_slots = [HOUR_TO_SLOT[h] for h in sorted(valid) if h in HOUR_TO_SLOT]
        return (
            f"Arrival slot '{arrival_slot}' doesn't match session '{session}'.\n"
            f"{SESSION_HINT[session]}\n"
            f"Valid slots for {session}: {', '.join(valid_slots)}"
        )
    return None


def build_features(patient, enc, feature_columns):
    gm        = enc["global_mean"]
    now       = datetime.today()
    day_abbr  = DAY_MAP.get(patient["day"], patient["day"])
    flow      = patient["flow_type"]
    specialty = patient["specialty"]
    session   = patient["session"]
    desig     = patient["consultant_desig"]
    arrival_num = enc["arrival_map"].get(patient["arrival_period"], 9.0)

    row = {
        "Age category":           patient["age_category"],
        "Gender":                 patient["gender"],
        "Patient visit category": patient["visit_category"],
        "Patient Flow type":      flow,
        "Specialty":              specialty,
        "Consultant Designation": desig,
        "Day":                    day_abbr,
        "Session":                session,
        "Arrival_Hour_Num":       arrival_num,
        "Is_Weekend":             1 if day_abbr in ["Sat", "Sun"] else 0,
        "Month":                  now.month,
        "Quarter":                math.ceil(now.month / 3),
        "Specialty_Volume":       enc.get("spec_volume", {}).get(specialty, 74522),
        "Flow_Session":           flow + "_" + session,
        "Specialty_Flow":         specialty + "_" + flow,
        "Session_Desig":          session + "_" + desig,
        "Specialty_Session":      specialty + "_" + session,
        "Flow_Desig":             flow + "_" + desig,
        "Specialty_AvgTime":      enc["spec_mean"].get(specialty, gm),
        "Flow_AvgTime":           enc["flow_mean"].get(flow, gm),
    }

    # Consultant Name — optional (bulk may omit)
    cons_name = patient.get("consultant_name", "")
    row["Consultant Name"] = cons_name if cons_name else ""
    row["Consultant_AvgTime"] = enc["consultant_mean"].get(cons_name, gm)

    df_row = pd.DataFrame([row])
    df_row = pd.get_dummies(df_row)
    df_row = df_row.reindex(columns=feature_columns, fill_value=0)
    return df_row


def run_predict(patient, payload):
    X        = build_features(patient, payload["encoding_maps"], payload["feature_columns"])
    pred_log = payload["model"].predict(X)[0]
    minutes  = float(np.expm1(pred_log))
    minutes  = max(5.0, min(300.0, minutes))
    low      = max(5,   round(minutes * 0.82))
    high     = min(300, round(minutes * 1.18))
    h, m     = divmod(round(minutes), 60)
    time_str = f"{h}h {m}m" if h else f"{m} min"
    return {"minutes": round(minutes, 1), "time_str": time_str, "low": low, "high": high}


def predict_bulk_df(df_in: pd.DataFrame, payload: dict) -> pd.DataFrame:
    """Run predictions on a normalised dataframe; return results df."""
    results = []
    enc     = payload["encoding_maps"]
    fcols   = payload["feature_columns"]
    now     = datetime.today()

    for idx, row in df_in.iterrows():
        session  = str(row.get("Session", "")).strip()
        arrival  = str(row.get("Arrival hour", "")).strip()
        err = validate_session_arrival(session, arrival)
        if err:
            results.append({
                "Row": idx + 2,  # +2: 1-indexed + header
                "Age_Category_Used": "",
                "Predicted_Min": None,
                "Low_Min": None,
                "High_Min": None,
                "Time_String": None,
                "Validation_Error": err.replace("\n", " | "),
            })
            continue

        # ── Resolve age category ──────────────────────────────────
        # Priority: numeric "Age" column > "Age category" column > default
        raw_age     = row.get("Age", None)
        raw_age_cat = str(row.get("Age category", "")).strip()

        if raw_age is not None and str(raw_age).strip() not in ("", "nan", "None"):
            # Numeric age → derive category automatically
            age_cat = age_to_category(raw_age)
        elif raw_age_cat:
            # Accept both "31-45 (Adult)" and "31-45"
            age_cat = raw_age_cat
            for long, short in AGE_LABEL_MAP.items():
                if age_cat == long:
                    age_cat = short
                    break
        else:
            age_cat = "31-45"  # safe fallback

        day_raw  = str(row.get("Day", "Monday")).strip()
        day_abbr = DAY_MAP.get(day_raw, day_raw)

        patient = {
            "age_category":     age_cat,
            "gender":           str(row.get("Gender", "Unknown")).strip(),
            "visit_category":   str(row.get("Patient visit category", "MRE")).strip(),
            "flow_type":        str(row.get("Patient Flow type", "Non-Dilated")).strip(),
            "specialty":        str(row.get("Specialty", "General")).strip(),
            "consultant_name":  "",
            "consultant_desig": str(row.get("Consultant Designation", "Specialist")).strip(),
            "day":              day_raw,
            "session":          session,
            "arrival_period":   arrival,
        }

        try:
            res = run_predict(patient, payload)
            results.append({
                "Row": idx + 2,
                "Age_Category_Used": age_cat,
                "Predicted_Min": res["minutes"],
                "Low_Min":       res["low"],
                "High_Min":      res["high"],
                "Time_String":   res["time_str"],
                "Validation_Error": "",
            })
        except Exception as ex:
            results.append({
                "Row": idx + 2,
                "Age_Category_Used": age_cat,
                "Predicted_Min": None,
                "Low_Min": None,
                "High_Min": None,
                "Time_String": None,
                "Validation_Error": str(ex),
            })

    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════
# GUI
# ══════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("EHR Workup Time Predictor")
        self.geometry("900x760")
        self.minsize(800, 620)
        self.configure(bg=BG)
        self.resizable(True, True)
        self.payload       = None
        self.bulk_result   = None
        self._pulse_job    = None
        self._anim_jobs    = []   # track all after() ids for cleanup
        self._btn_anim     = {}   # per-button animation state
        self._build_fonts()
        self._build_styles()
        self._build_ui()
        self._try_load_model_async()

    # ── fonts / styles ────────────────────────────────────────────
    def _build_fonts(self):
        self.f_title  = tkfont.Font(family="Segoe UI",    size=16, weight="bold")
        self.f_head   = tkfont.Font(family="Segoe UI",    size=10, weight="bold")
        self.f_body   = tkfont.Font(family="Segoe UI",    size=10)
        self.f_big    = tkfont.Font(family="Segoe UI",    size=34, weight="bold")
        self.f_sub    = tkfont.Font(family="Segoe UI",    size=11)
        self.f_small  = tkfont.Font(family="Segoe UI",    size=8)
        self.f_badge  = tkfont.Font(family="Segoe UI",    size=8,  weight="bold")
        self.f_mono   = tkfont.Font(family="Consolas",    size=10)
        self.f_number = tkfont.Font(family="Segoe UI",    size=22, weight="bold")

    def _build_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("TCombobox",
            fieldbackground=CARD2, background=CARD2, foreground=TEXT,
            selectbackground=ACCENT, selectforeground=WHITE,
            bordercolor=BORDER2, lightcolor=BORDER2, darkcolor=BORDER2,
            arrowcolor=TEXT_DIM, padding=7)
        s.map("TCombobox",
            fieldbackground=[("readonly", CARD2)],
            foreground=[("readonly", TEXT)],
            bordercolor=[("focus", ACCENT)])
        s.configure("Vertical.TScrollbar",
            background=BORDER, troughcolor=PANEL,
            bordercolor=PANEL, arrowcolor=TEXT_DIM, width=6)
        s.configure("TNotebook",
            background=PANEL, borderwidth=0, tabmargins=0)
        s.configure("TNotebook.Tab",
            background=PANEL, foreground=TEXT_MUTED,
            padding=[20, 10], font=("Segoe UI", 9, "bold"),
            borderwidth=0)
        s.map("TNotebook.Tab",
            background=[("selected", BG)],
            foreground=[("selected", TEXT)])
        s.configure("Treeview",
            background=CARD, fieldbackground=CARD,
            foreground=TEXT, rowheight=28,
            borderwidth=0, relief="flat", font=("Segoe UI", 9))
        s.configure("Treeview.Heading",
            background=PANEL, foreground=TEXT_DIM,
            font=("Segoe UI", 8, "bold"), borderwidth=0, relief="flat")
        s.map("Treeview", background=[("selected", ACCENT_GLO)])
        s.configure("Accent.TProgressbar",
            troughcolor=BORDER, background=ACCENT,
            borderwidth=0, thickness=3)

    # ── animation helpers ─────────────────────────────────────────
    def _animate_bar(self, canvas, bar_id, target_pct, colour, step=0):
        """Smoothly animate a canvas bar from 0 → target_pct (ease-out cubic)."""
        total = 40
        if step > total:
            return
        ease  = target_pct * (1 - (1 - step / total) ** 3)
        w     = canvas.winfo_width() or 500
        canvas.coords(bar_id, 0, 0, max(1, int(w * ease / 100)), 10)
        canvas.itemconfig(bar_id, fill=colour)
        if step < total:
            self.after(14, lambda: self._animate_bar(canvas, bar_id, target_pct, colour, step + 1))

    def _animate_count(self, label, target, unit="", step=0, total=30):
        """Count-up animation with ease-out."""
        if step > total:
            label.configure(text=f"{target}{unit}")
            return
        ease = int(target * (1 - (1 - step / total) ** 2))
        label.configure(text=f"{ease}{unit}")
        self.after(18, lambda: self._animate_count(label, target, unit, step + 1, total))

    def _pulse_dot(self, canvas, dot_id, phase=0):
        """Continuously pulsing dot on header."""
        bright = ["#3D8EF5", "#5BA3FF", "#7AB8FF", "#5BA3FF", "#3D8EF5",
                  "#2A6ACC", "#1E52A8", "#2A6ACC"]
        colour = bright[int(phase) % len(bright)]
        canvas.itemconfig(dot_id, fill=colour)
        self._pulse_job = self.after(80, lambda: self._pulse_dot(canvas, dot_id, phase + 1))

    def _fade_in(self, widget, alpha=0, steps=16):
        """Fade a label from dim → full TEXT colour."""
        if alpha >= steps:
            try: widget.configure(fg=TEXT)
            except: pass
            return
        ratio = alpha / steps
        r = int(0x37 + (0xE8 - 0x37) * ratio)
        g = int(0x40 + (0xED - 0x40) * ratio)
        b = int(0x60 + (0xF8 - 0x60) * ratio)
        try:
            widget.configure(fg=f"#{r:02x}{g:02x}{b:02x}")
        except: return
        self.after(16, lambda: self._fade_in(widget, alpha + 1, steps))

    def _slide_in(self, widget, start_y, end_y=0, step=0, total=18):
        """Slide a widget down from start_y offset toward end_y using place."""
        if step > total:
            try: widget.place_forget()
            except: pass
            return
        ease = start_y + (end_y - start_y) * (1 - (1 - step / total) ** 3)
        try:
            widget.place(x=0, y=int(ease), relwidth=1)
        except: return
        self.after(14, lambda: self._slide_in(widget, start_y, end_y, step + 1, total))

    def _shimmer_btn(self, btn, normal, hover, pressing=False):
        """Brief flash on button click."""
        flash_cols = [hover, "#8B6FFF", hover, normal] if pressing else [hover, normal]
        def _step(i=0):
            if i >= len(flash_cols): return
            try: btn.configure(bg=flash_cols[i])
            except: return
            self.after(60, lambda: _step(i + 1))
        _step()

    def _animated_btn_hover(self, btn, normal, hover):
        """Smooth colour interpolation on hover (6 steps)."""
        def _interp(c1, c2, t):
            r = int(int(c1[1:3], 16) + (int(c2[1:3], 16) - int(c1[1:3], 16)) * t)
            g = int(int(c1[3:5], 16) + (int(c2[3:5], 16) - int(c1[3:5], 16)) * t)
            b = int(int(c1[5:7], 16) + (int(c2[5:7], 16) - int(c1[5:7], 16)) * t)
            return f"#{r:02x}{g:02x}{b:02x}"
        steps = 6
        key   = id(btn)

        def _run(target, step=0):
            if self._btn_anim.get(key) != target: return
            t = step / steps
            col = _interp(normal, hover, t) if target == "in" else _interp(hover, normal, t)
            try: btn.configure(bg=col)
            except: return
            if step < steps:
                self.after(20, lambda: _run(target, step + 1))

        btn.bind("<Enter>",  lambda e: [self._btn_anim.update({key: "in"}),  _run("in")])
        btn.bind("<Leave>",  lambda e: [self._btn_anim.update({key: "out"}), _run("out")])
        btn.bind("<Button-1>", lambda e: self._shimmer_btn(btn, normal, hover, True))

    # ── main layout ───────────────────────────────────────────────
    def _build_ui(self):
        # ── Animated header ────────────────────────────────────────
        hdr_canvas = tk.Canvas(self, height=72, bg=PANEL,
                               highlightthickness=0, bd=0)
        hdr_canvas.pack(fill="x", side="top")

        # Gradient stripes (dark top → darker bottom)
        for i in range(72):
            t = i / 72
            r = int(0x0E * (1 - t) + 0x07 * t)
            g = int(0x14 * (1 - t) + 0x0B * t)
            b = int(0x20 * (1 - t) + 0x14 * t)
            hdr_canvas.create_line(0, i, 3000, i, fill=f"#{r:02x}{g:02x}{b:02x}")

        # Blue accent left stripe
        hdr_canvas.create_rectangle(0, 0, 4, 72, fill=ACCENT, outline="")

        # Subtle right-side glow blob
        for gx in range(20):
            alpha = int(15 * (1 - gx / 20))
            col   = f"#{alpha:02x}{alpha*2:02x}{alpha*4:02x}"
            hdr_canvas.create_rectangle(
                3000 - gx * 18, 0, 3000, 72, fill=col, outline="")

        # Pulsing dot
        self._hdr_dot = hdr_canvas.create_oval(24, 30, 38, 44, fill=ACCENT, outline="")
        self._pulse_dot(hdr_canvas, self._hdr_dot)

        # Title
        hdr_canvas.create_text(52, 36, text="Workup Time Predictor",
            font=self.f_title, fill=TEXT, anchor="w")

        # Subtitle
        hdr_canvas.create_text(52, 56, text="EHR Prediction System  ·  XGBoost",
            font=self.f_small, fill=TEXT_MUTED, anchor="w")

        # Status label
        self.status_lbl = tk.Label(self, text="● Loading model…",
                                   font=self.f_small, bg=PANEL, fg=WARN)
        self.status_lbl.place(in_=hdr_canvas, relx=1.0, rely=0.5,
                               anchor="e", x=-24)

        # Thin separator below header
        tk.Frame(self, bg=BORDER2, height=1).pack(fill="x")

        # ── Notebook ───────────────────────────────────────────────
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)

        tab_single = tk.Frame(self.notebook, bg=BG)
        tab_bulk   = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(tab_single, text="  ⊙  Single Patient  ")
        self.notebook.add(tab_bulk,   text="  ⊞  Bulk Predict  ")

        self._build_single_tab(tab_single)
        self._build_bulk_tab(tab_bulk)

    # ══════════════════════════════════════════════════════════════
    # SINGLE PATIENT TAB
    # ══════════════════════════════════════════════════════════════
    def _build_single_tab(self, parent):
        outer  = tk.Frame(parent, bg=BG)
        outer.pack(fill="both", expand=True)

        canvas    = tk.Canvas(outer, bg=BG, highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview,
                                  style="Vertical.TScrollbar")
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self.scroll_frame = tk.Frame(canvas, bg=BG)
        win = canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")

        self.scroll_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(win, width=e.width))
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        self._build_form(self.scroll_frame)

    def _build_form(self, parent):
        # PATIENT DEMOGRAPHICS
        self._section(parent, "PATIENT DEMOGRAPHICS")
        row0 = tk.Frame(parent, bg=BG)
        row0.pack(fill="x", padx=10, pady=(0, 12))
        self.var_gender = self._combo(row0, "Gender",
            ["Male", "Female", "Unknown"], side="left", width=16)
        tk.Frame(row0, bg=BG, width=16).pack(side="left")
        self.var_age_cat = self._combo(row0, "Age Group",
            list(AGE_LABEL_MAP.keys()), side="left", width=28)

        # VISIT DETAILS
        self._section(parent, "VISIT DETAILS")
        row1 = tk.Frame(parent, bg=BG)
        row1.pack(fill="x", padx=10, pady=(0, 12))
        self.var_visit_cat = self._combo(row1, "Visit Category",
            ["MRE", "REG", "SRE"], side="left", width=12)
        tk.Frame(row1, bg=BG, width=16).pack(side="left")
        self.var_flow = self._combo(row1, "Patient Flow Type",
            ["Non-Dilated", "Dilated", "Procedure"], side="left", width=18)
        tk.Frame(row1, bg=BG, width=16).pack(side="left")
        self.var_desig = self._combo(row1, "Consultant Designation",
            ["Specialist", "Post Graduate"], side="left", width=18)

        row2 = tk.Frame(parent, bg=BG)
        row2.pack(fill="x", padx=10, pady=(0, 12))
        self.var_specialty = self._combo(row2, "Specialty",
            ["Retina", "General", "Glaucoma", "Cornea", "Pediatric and Low vision"],
            side="left", width=34)

        # SCHEDULE
        self._section(parent, "SCHEDULE")
        row3 = tk.Frame(parent, bg=BG)
        row3.pack(fill="x", padx=10, pady=(0, 4))

        self.var_day = self._combo(row3, "Day of Visit",
            ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"],
            side="left", width=14)
        tk.Frame(row3, bg=BG, width=16).pack(side="left")

        # Arrival slot only — session auto-derived
        arrival_wrap = tk.Frame(row3, bg=BG)
        arrival_wrap.pack(side="left")
        tk.Label(arrival_wrap, text="Arrival Slot", font=self.f_small,
                 bg=BG, fg=TEXT_DIM).pack(anchor="w")
        self.var_arrival = tk.StringVar(value=list(ARRIVAL_MAP.keys())[0])
        self._arrival_cb = ttk.Combobox(arrival_wrap, textvariable=self.var_arrival,
            values=list(ARRIVAL_MAP.keys()),
            state="readonly", width=22, font=self.f_body)
        self._arrival_cb.pack(anchor="w", pady=(2, 0))

        # Auto-session badge row
        badge_row = tk.Frame(parent, bg=BG)
        badge_row.pack(anchor="w", padx=24, pady=(8, 4))
        tk.Label(badge_row, text="Session auto-detected: ",
                 font=self.f_small, bg=BG, fg=TEXT_DIM).pack(side="left")
        self._session_badge = tk.Label(badge_row, text="",
            font=self.f_badge, bg="#2E6B4F", fg=WHITE, padx=8, pady=3)
        self._session_badge.pack(side="left")
        self._session_hint_lbl = tk.Label(badge_row, text="",
            font=self.f_small, bg=BG, fg=TEXT_DIM)
        self._session_hint_lbl.pack(side="left", padx=(8, 0))
        self._update_session_badge()

        # bind arrival to update badge
        self.var_arrival.trace_add("write", lambda *_: self._update_session_badge())

        # PREDICT BUTTON
        tk.Frame(parent, bg=BG, height=8).pack()
        btn_frame = tk.Frame(parent, bg=BG)
        btn_frame.pack(fill="x", padx=24, pady=(0, 20))
        self.btn = tk.Button(btn_frame, text="PREDICT WORKUP TIME",
            font=self.f_head, bg=ACCENT, fg=WHITE, relief="flat",
            activebackground=ACCENT2, activeforeground=WHITE,
            cursor="hand2", bd=0, padx=0, pady=14,
            command=self._on_predict)
        self.btn.pack(fill="x")
        self._btn_hover(self.btn, ACCENT, ACCENT2)

        # RESULT AREA
        self.result_outer = tk.Frame(parent, bg=BG)
        self.result_outer.pack(fill="x", padx=24, pady=(0, 32))
        self._build_result_placeholder()

    def _update_session_badge(self):
        session = arrival_to_session(self.var_arrival.get())
        color   = SESSION_BADGE_COLOR.get(session, ACCENT)
        hints   = {
            "Forenoon":  "  08:00 – 13:00  (before 1 PM)",
            "Lunch":     "  13:00 – 15:00  (1 PM – 3 PM)",
            "Afternoon": "  15:00 – 17:00  (after 3 PM)",
        }
        self._session_badge.configure(text=f"  {session}  ", bg=color)
        self._session_hint_lbl.configure(text=hints.get(session, ""))

    def _section(self, parent, text, icon=""):
        f = tk.Frame(parent, bg=BG)
        f.pack(fill="x", padx=20, pady=(18, 6))
        # accent line + label
        tk.Frame(f, bg=ACCENT, width=3, height=16).pack(side="left")
        tk.Label(f, text=f"  {icon}{text}", font=self.f_badge,
                 bg=BG, fg=TEXT_DIM).pack(side="left", pady=(0, 0))
        tk.Frame(f, bg=BORDER, height=1).pack(side="bottom", fill="x")

    def _combo(self, parent, label, values, side="left", width=20):
        wrap = tk.Frame(parent, bg=BG)
        wrap.pack(side=side, padx=(0, 0))
        tk.Label(wrap, text=label, font=self.f_small,
                 bg=BG, fg=TEXT_DIM).pack(anchor="w", pady=(0, 2))
        var = tk.StringVar(value=values[0])
        cb  = ttk.Combobox(wrap, textvariable=var, values=values,
                            state="readonly", width=width, font=self.f_body)
        cb.pack(anchor="w")
        # glow border on focus using canvas underline
        under = tk.Frame(wrap, bg=BORDER, height=2)
        under.pack(fill="x")
        cb.bind("<FocusIn>",  lambda e: under.configure(bg=ACCENT))
        cb.bind("<FocusOut>", lambda e: under.configure(bg=BORDER))
        return var

    def _btn_hover(self, btn, normal, hover):
        self._animated_btn_hover(btn, normal, hover)

    def _build_result_placeholder(self):
        for w in self.result_outer.winfo_children():
            w.destroy()
        tk.Label(self.result_outer,
                 text="Fill in the form above and click Predict",
                 font=self.f_small, bg=BG, fg=TEXT_MUTED).pack(pady=12)

    def _show_result(self, res):
        for w in self.result_outer.winfo_children():
            w.destroy()

        minutes   = res["minutes"]
        low, high = res["low"], res["high"]
        pct       = min(100, int((minutes / 240) * 100))
        colour    = SUCCESS if minutes <= 60 else WARN if minutes <= 120 else DANGER
        h, m      = divmod(round(minutes), 60)

        # Outer card with left coloured border accent
        card = tk.Frame(self.result_outer, bg=CARD,
                        highlightbackground=BORDER2, highlightthickness=1)
        card.pack(fill="x")

        # Left accent stripe animated colour
        stripe = tk.Frame(card, bg=colour, width=4)
        stripe.pack(side="left", fill="y")

        inner = tk.Frame(card, bg=CARD)
        inner.pack(side="left", fill="both", expand=True)

        # ── Top row: big time + mini stat cards ───────────────────
        top = tk.Frame(inner, bg=CARD)
        top.pack(fill="x", padx=20, pady=(18, 10))

        left_col = tk.Frame(top, bg=CARD)
        left_col.pack(side="left")

        tk.Label(left_col, text="PREDICTED WORKUP TIME",
                 font=self.f_badge, bg=CARD, fg=TEXT_MUTED).pack(anchor="w")

        # Animated count-up for minutes
        time_lbl = tk.Label(left_col, text="0 min",
                            font=self.f_big, bg=CARD, fg=colour)
        time_lbl.pack(anchor="w")
        self._fade_in(time_lbl)
        display_str = f"{h}h {m}m" if h else f"{m} min"
        self._animate_count(time_lbl, round(minutes),
                            unit=(" min" if not h else ""),
                            total=28)
        # Correct to h/m format after count finishes
        self.after(600, lambda: time_lbl.configure(text=display_str))

        conf_lbl = tk.Label(left_col,
                            text=f"Confidence range  {low} – {high} min",
                            font=self.f_sub, bg=CARD, fg=TEXT_DIM)
        conf_lbl.pack(anchor="w", pady=(4, 0))
        self._fade_in(conf_lbl)

        right_col = tk.Frame(top, bg=CARD)
        right_col.pack(side="right")

        for val, lbl, col in [
            (f"{low} min",          "Optimistic",    SUCCESS_DIM),
            (f"{round(minutes)} min","Predicted",    CARD2),
            (f"{high} min",         "Conservative",  WARN_DIM),
        ]:
            s = tk.Frame(right_col, bg=col,
                         highlightbackground=BORDER2, highlightthickness=1)
            s.pack(side="left", padx=3, ipadx=2)
            vl = tk.Label(s, text=val, font=self.f_sub, bg=col, fg=TEXT, width=10)
            vl.pack(pady=(10, 2), padx=8)
            tk.Label(s, text=lbl, font=self.f_small, bg=col, fg=TEXT_DIM).pack(pady=(0, 10), padx=8)
            self._fade_in(vl)

        # ── Animated progress bar ─────────────────────────────────
        bar_frame = tk.Frame(inner, bg=CARD)
        bar_frame.pack(fill="x", padx=20, pady=(0, 10))

        tk.Label(bar_frame, text="TIME PROFILE  (0 – 240 min)",
                 font=self.f_small, bg=CARD, fg=TEXT_MUTED).pack(anchor="w", pady=(0, 6))

        track_bg = tk.Canvas(bar_frame, height=10, bg=BORDER,
                             highlightthickness=0, bd=0)
        track_bg.pack(fill="x")
        # Segment markers at 60, 120, 180
        def _draw_markers(event=None):
            w = track_bg.winfo_width()
            if w < 10: return
            for frac in [0.25, 0.5, 0.75]:
                x = int(w * frac)
                track_bg.create_line(x, 0, x, 10, fill=BORDER2, width=1)
        track_bg.bind("<Configure>", _draw_markers)
        track_bg.update_idletasks()
        _draw_markers()

        bar_id = track_bg.create_rectangle(0, 0, 0, 10, fill=colour, outline="")
        self.after(120, lambda: self._animate_bar(track_bg, bar_id, pct, colour))

        tick_row = tk.Frame(bar_frame, bg=CARD)
        tick_row.pack(fill="x", pady=(2, 0))
        for tick in ["0", "60", "120", "180", "240 min"]:
            tk.Label(tick_row, text=tick, font=self.f_small,
                     bg=CARD, fg=TEXT_MUTED).pack(side="left", expand=True)

        # ── Tag pills ─────────────────────────────────────────────
        tag_frame = tk.Frame(inner, bg=CARD)
        tag_frame.pack(fill="x", padx=20, pady=(6, 18))

        age_data_val = AGE_LABEL_MAP.get(self.var_age_cat.get(), "")
        session_now  = arrival_to_session(self.var_arrival.get())
        badge_col    = {
            "Forenoon":  "#1A4A35", "Lunch": "#4A3A10", "Afternoon": "#0E2E5A"
        }
        tags = [
            (self.var_flow.get(),      ACCENT),
            (self.var_specialty.get(), ACCENT2),
            (session_now,              badge_col.get(session_now, ACCENT)),
            (age_data_val,             "#303850"),
        ]
        if self.var_day.get() in ["Saturday", "Sunday"]:
            tags.append(("Weekend", WARN_DIM))
        if self.var_visit_cat.get() == "REG":
            tags.append(("New Registration", DANGER_DIM))

        for i, (txt, col) in enumerate(tags):
            if not txt: continue
            pill = tk.Label(tag_frame, text=f"  {txt}  ",
                            font=self.f_badge, bg=col, fg=TEXT,
                            padx=4, pady=4)
            pill.pack(side="left", padx=(0, 6))
            # staggered fade
            self.after(i * 60, lambda w=pill: self._fade_in(w))

    def _on_predict(self):
        if self.payload is None:
            messagebox.showwarning("Model not loaded",
                "workup_model_age7_ar1.pkl not found.\n"
                "Run train_age7_ar1.py first.")
            return

        session = arrival_to_session(self.var_arrival.get())
        age_cat = AGE_LABEL_MAP.get(self.var_age_cat.get(), "31-45")
        patient = {
            "gender":           self.var_gender.get(),
            "age_category":     age_cat,
            "visit_category":   self.var_visit_cat.get(),
            "flow_type":        self.var_flow.get(),
            "specialty":        self.var_specialty.get(),
            "consultant_name":  "",
            "consultant_desig": self.var_desig.get(),
            "day":              self.var_day.get(),
            "session":          session,
            "arrival_period":   self.var_arrival.get(),
        }

        self.btn.configure(state="disabled", text="Predicting…")
        self.update_idletasks()

        def worker():
            try:
                res = run_predict(patient, self.payload)
                self.after(0, lambda: self._show_result(res))
            except Exception as ex:
                self.after(0, lambda: messagebox.showerror("Prediction error", str(ex)))
            finally:
                self.after(0, lambda: self.btn.configure(
                    state="normal", text="PREDICT WORKUP TIME"))

        threading.Thread(target=worker, daemon=True).start()

    # ══════════════════════════════════════════════════════════════
    # BULK PREDICT TAB
    # ══════════════════════════════════════════════════════════════
    def _build_bulk_tab(self, parent):
        # ── Instructions card ─────────────────────────────────────
        info = tk.Frame(parent, bg=CARD,
                        highlightbackground=BORDER, highlightthickness=1)
        info.pack(fill="x", padx=24, pady=(20, 0))

        tk.Label(info, text="📂  Bulk Prediction from Excel",
                 font=self.f_head, bg=CARD, fg=TEXT).pack(anchor="w", padx=16, pady=(14, 4))

        instructions = (
            "Upload an .xlsx file with one patient per row.\n"
            "Required columns (any order, case-insensitive):\n"
            "  Gender · Patient visit category · Patient Flow type\n"
            "  Specialty · Consultant Designation · Day · Session · Arrival hour\n"
            "Age (pick ONE):\n"
            "  • 'Age'          → numeric patient age (e.g. 14) — auto-mapped to category\n"
            "  • 'Age category' → direct bin (e.g. 13-19 or 13-19 (Teenage))\n\n"
            "Session rules enforced per row:\n"
            "  Forenoon → 08:00–13:00 · Lunch → 13:00–15:00 · Afternoon → 15:00–17:00\n"
            "Rows with mismatches will show an error — not skipped.\n"
            "Results are auto-saved to Excel after prediction."
        )
        tk.Label(info, text=instructions, font=self.f_small,
                 bg=CARD, fg=TEXT_DIM, justify="left").pack(anchor="w", padx=16, pady=(0, 14))

        # ── File picker row ───────────────────────────────────────
        pick_row = tk.Frame(parent, bg=BG)
        pick_row.pack(fill="x", padx=24, pady=(16, 0))

        self._bulk_path_var = tk.StringVar(value="No file selected")
        tk.Label(pick_row, textvariable=self._bulk_path_var, font=self.f_small,
                 bg=BG, fg=TEXT_DIM, anchor="w").pack(side="left", fill="x", expand=True)

        browse_btn = tk.Button(pick_row, text="Browse Excel…",
            font=self.f_body, bg=PANEL, fg=TEXT, relief="flat",
            activebackground=BORDER, activeforeground=WHITE,
            cursor="hand2", bd=0, padx=14, pady=8,
            command=self._bulk_browse)
        browse_btn.pack(side="right")
        self._btn_hover(browse_btn, PANEL, BORDER)

        # ── Run button ────────────────────────────────────────────
        run_row = tk.Frame(parent, bg=BG)
        run_row.pack(fill="x", padx=24, pady=(10, 0))

        self.bulk_btn = tk.Button(run_row, text="RUN BULK PREDICTION",
            font=self.f_head, bg=ACCENT, fg=WHITE, relief="flat",
            activebackground=ACCENT2, activeforeground=WHITE,
            cursor="hand2", bd=0, padx=0, pady=14,
            command=self._on_bulk_predict)
        self.bulk_btn.pack(fill="x")
        self._btn_hover(self.bulk_btn, ACCENT, ACCENT2)

        # ── Action row: progress label + save button ─────────────
        action_row = tk.Frame(parent, bg=BG)
        action_row.pack(fill="x", padx=24, pady=(10, 0))

        self.bulk_status = tk.Label(action_row, text="",
            font=self.f_small, bg=BG, fg=TEXT_DIM, anchor="w")
        self.bulk_status.pack(side="left", fill="x", expand=True)

        # One-click save — auto-saves to same folder as input, no dialog
        self.export_btn = tk.Button(action_row,
            text="💾  Save to Excel",
            font=self.f_body, bg=SUCCESS_DIM, fg=SUCCESS, relief="flat",
            activebackground="#1A5A35", activeforeground=SUCCESS,
            cursor="hand2", bd=0, padx=16, pady=8,
            state="disabled",
            command=self._bulk_export_auto_save)
        self.export_btn.pack(side="right")
        self._animated_btn_hover(self.export_btn, SUCCESS_DIM, "#1A5A35")

        # ── Results area ──────────────────────────────────────────
        results_frame = tk.Frame(parent, bg=BG)
        results_frame.pack(fill="both", expand=True, padx=24, pady=(12, 0))

        # Summary bar (hidden until results arrive)
        self.bulk_summary_bar = tk.Frame(results_frame, bg=CARD,
                                          highlightbackground=BORDER, highlightthickness=1)

        # Treeview
        tree_scroll_y = ttk.Scrollbar(results_frame, orient="vertical",
                                       style="Vertical.TScrollbar")
        tree_scroll_x = ttk.Scrollbar(results_frame, orient="horizontal")

        cols = ("Row", "Age Cat", "Predicted", "Low", "High", "Status")
        self.bulk_tree = ttk.Treeview(results_frame, columns=cols,
                                       show="headings",
                                       yscrollcommand=tree_scroll_y.set,
                                       xscrollcommand=tree_scroll_x.set)
        tree_scroll_y.configure(command=self.bulk_tree.yview)
        tree_scroll_x.configure(command=self.bulk_tree.xview)

        col_widths = {"Row": 50, "Age Cat": 90, "Predicted": 110, "Low": 80, "High": 80, "Status": 360}
        for c in cols:
            self.bulk_tree.heading(c, text=c)
            self.bulk_tree.column(c, width=col_widths[c], anchor="center")
        self.bulk_tree.column("Status", anchor="w")

        # Tag colours for tree rows
        self.bulk_tree.tag_configure("ok",    foreground=SUCCESS)
        self.bulk_tree.tag_configure("error", foreground=DANGER)

        tree_scroll_y.pack(side="right",  fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        self.bulk_tree.pack(fill="both", expand=True)

        # store ref for summary labels
        self._bulk_summary_labels = {}

    def _bulk_browse(self):
        path = filedialog.askopenfilename(
            title="Select patient Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self._bulk_file_path = path
            self._bulk_path_var.set(os.path.basename(path))
            self.bulk_status.configure(text=f"File loaded: {os.path.basename(path)}", fg=TEXT_DIM)

    def _on_bulk_predict(self):
        if self.payload is None:
            messagebox.showwarning("Model not loaded",
                "workup_model_age7_ar1.pkl not found.\n"
                "Run train_age7_ar1.py first.")
            return

        path = getattr(self, "_bulk_file_path", None)
        if not path:
            messagebox.showwarning("No file", "Please browse and select an Excel file first.")
            return

        self.bulk_btn.configure(state="disabled", text="Running…")
        self.bulk_status.configure(text="Reading file…", fg=WARN)
        self.export_btn.configure(state="disabled", text="💾  Save to Excel")

        # Animated dots on status while running
        def _spin(dots=0):
            if not hasattr(self, "_bulk_spinning") or not self._bulk_spinning: return
            self.bulk_status.configure(
                text="Predicting" + "." * (dots % 4), fg=WARN)
            self.after(350, lambda: _spin(dots + 1))
        self._bulk_spinning = True
        self.after(400, _spin)

        # Clear previous results
        for row in self.bulk_tree.get_children():
            self.bulk_tree.delete(row)
        self.bulk_summary_bar.pack_forget()

        def worker():
            try:
                df_raw = pd.read_excel(path, header=0)
                # Normalise column names
                df_raw.columns = [c.strip() for c in df_raw.columns]
                rename = {}
                for col in df_raw.columns:
                    mapped = BULK_COL_MAP.get(col.lower())
                    if mapped:
                        rename[col] = mapped
                df_raw.rename(columns=rename, inplace=True)

                # Check required columns
                missing = [c for c in REQUIRED_BULK_COLS if c not in df_raw.columns]
                # Age: need at least "Age" or "Age category"
                if not any(c in df_raw.columns for c in AGE_COLS_EITHER):
                    missing.append("Age (or Age category)")
                if missing:
                    raise ValueError(
                        f"Missing required columns:\n  {', '.join(missing)}\n\n"
                        "Check the column names in your Excel file match the expected names."
                    )

                n = len(df_raw)
                self.after(0, lambda: self.bulk_status.configure(
                    text=f"Predicting {n:,} rows…", fg=WARN))

                result_df = predict_bulk_df(df_raw, self.payload)
                self.bulk_result = result_df

                # Merge original data + results for export
                self._bulk_export_df = pd.concat(
                    [df_raw.reset_index(drop=True), result_df.reset_index(drop=True)], axis=1)

                self.after(0, lambda: self._show_bulk_results(result_df, n))

            except Exception as ex:
                self.after(0, lambda: messagebox.showerror("Bulk Predict Error", str(ex)))
                self.after(0, lambda: self.bulk_status.configure(
                    text="Error — see dialog", fg=DANGER))
            finally:
                self._bulk_spinning = False
                self.after(0, lambda: self.bulk_btn.configure(
                    state="normal", text="RUN BULK PREDICTION"))

        threading.Thread(target=worker, daemon=True).start()

    def _show_bulk_results(self, result_df: pd.DataFrame, total: int):
        n_ok  = result_df["Validation_Error"].eq("").sum()
        n_err = total - n_ok

        # ── Animated summary bar ──────────────────────────────────
        self.bulk_summary_bar.pack(fill="x", pady=(8, 6))
        for w in self.bulk_summary_bar.winfo_children():
            w.destroy()

        avg = 0
        if n_ok > 0:
            avg = int(result_df.loc[result_df["Validation_Error"] == "",
                                    "Predicted_Min"].mean())

        summary_items = [
            (total,   "Total Rows",  TEXT,    ""),
            (n_ok,    "Predicted",   SUCCESS, ""),
            (n_err,   "Errors",      DANGER if n_err else TEXT_MUTED, ""),
            (avg,     "Avg Workup",  WARN,    " min"),
        ]

        for i, (val, lbl, col, unit) in enumerate(summary_items):
            cell = tk.Frame(self.bulk_summary_bar, bg=CARD2,
                            highlightbackground=BORDER2, highlightthickness=1)
            cell.pack(side="left", padx=(0 if i == 0 else 6, 0),
                      pady=8, ipadx=14, ipady=6)
            num_lbl = tk.Label(cell, text="0", font=self.f_sub, bg=CARD2, fg=col)
            num_lbl.pack(pady=(8, 2), padx=12)
            tk.Label(cell, text=lbl, font=self.f_small, bg=CARD2,
                     fg=TEXT_DIM).pack(pady=(0, 8), padx=12)
            # staggered count-up
            self.after(i * 80,
                lambda w=num_lbl, v=val, u=unit: self._animate_count(w, v, unit=u))

        # ── Populate treeview (rows fade in progressively) ────────
        def _insert_rows(rows, idx=0, batch=8):
            if idx >= len(rows): return
            for row in rows[idx:idx + batch]:
                is_ok = row["Validation_Error"] == ""
                tag   = "ok" if is_ok else "error"
                self.bulk_tree.insert("", "end",
                    values=(
                        int(row["Row"]),
                        row["Age_Category_Used"] if is_ok else "—",
                        f"{row['Predicted_Min']:.0f} min" if is_ok else "—",
                        f"{row['Low_Min']:.0f}"           if is_ok else "—",
                        f"{row['High_Min']:.0f}"          if is_ok else "—",
                        row["Time_String"] if is_ok
                            else f"⚠  {row['Validation_Error']}",
                    ),
                    tags=(tag,))
            if idx + batch < len(rows):
                self.after(30, lambda: _insert_rows(rows, idx + batch))

        rows = result_df.to_dict("records")
        _insert_rows(rows)

        status_txt = f"✓  Done — {n_ok:,} predicted, {n_err:,} error(s)."
        status_col = SUCCESS if n_err == 0 else WARN
        self.bulk_status.configure(text=status_txt, fg=status_col)
        self._fade_in(self.bulk_status)
        self.export_btn.configure(state="normal")

    def _bulk_export_auto_save(self):
        """One-click save — writes to same folder as the input file, no dialog."""
        df = getattr(self, "_bulk_export_df", None)
        if df is None:
            return
        src_path = getattr(self, "_bulk_file_path", None)
        if src_path:
            folder   = os.path.dirname(os.path.abspath(src_path))
            stem     = os.path.splitext(os.path.basename(src_path))[0]
            save_path = os.path.join(folder, f"{stem}_predictions.xlsx")
        else:
            save_path = os.path.join(os.getcwd(), "workup_predictions.xlsx")

        self.export_btn.configure(state="disabled", text="Saving…")
        self.update_idletasks()

        def worker():
            try:
                self._write_styled_excel(df, save_path)
                self.after(0, lambda: self.bulk_status.configure(
                    text=f"✓  Saved → {os.path.basename(save_path)}  ({os.path.dirname(save_path)})",
                    fg=SUCCESS))
                self.after(0, lambda: self._fade_in(self.bulk_status))
            except Exception as ex:
                self.after(0, lambda: messagebox.showerror("Save failed", str(ex)))
            finally:
                self.after(0, lambda: self.export_btn.configure(
                    state="normal", text="💾  Save to Excel"))

        threading.Thread(target=worker, daemon=True).start()

    def _bulk_export(self):
        """Legacy method kept for compatibility."""
        self._bulk_export_auto_save()

    def _write_styled_excel(self, df: pd.DataFrame, save_path: str):
        try:
            from openpyxl import load_workbook as _lw
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter

            df.to_excel(save_path, index=False)
            wb = _lw(save_path)
            ws = wb.active

            thin    = Side(style="thin", color="2A2F3F")
            border  = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_fill = PatternFill("solid", start_color="1E2330")
            hdr_font = Font(bold=True, color="F0F2F8", name="Arial", size=10)
            ok_font  = Font(name="Arial", size=10, color="3DDC84")
            err_font = Font(name="Arial", size=10, color="F75F5F")
            nor_font = Font(name="Arial", size=10, color="C8CDDC")
            fill_odd  = PatternFill("solid", start_color="1E2330")
            fill_even = PatternFill("solid", start_color="181C27")

            # result columns (added by predict)
            result_cols = {"Row", "Age_Category_Used", "Predicted_Min",
                           "Low_Min", "High_Min", "Time_String", "Validation_Error"}
            col_names = [c.value for c in ws[1]]

            for cell in ws[1]:
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                fill = fill_odd if r_idx % 2 else fill_even
                # check if this row has an error
                err_val = ""
                for cell in row:
                    if col_names[cell.column - 1] == "Validation_Error":
                        err_val = str(cell.value or "")
                for cell in row:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                    cname = col_names[cell.column - 1]
                    if cname in result_cols:
                        cell.font = err_font if err_val else ok_font
                    else:
                        cell.font = nor_font

            # Auto-width
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)

            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"
            wb.save(save_path)
            # caller handles success feedback
        except Exception as ex:
            raise  # re-raise so caller can catch

    # ── Model loader ──────────────────────────────────────────────
    def _try_load_model_async(self):
        def worker():
            try:
                self.payload = load_model("workup_model_age7_ar1.pkl")
                self.after(0, lambda: self.status_lbl.configure(
                    text="● Model ready", fg=SUCCESS))
            except FileNotFoundError:
                self.after(0, lambda: self.status_lbl.configure(
                    text="⚠ workup_model_age7_ar1.pkl not found — train first", fg=WARN))
            except Exception as ex:
                self.after(0, lambda: self.status_lbl.configure(
                    text=f"⚠ Model error: {ex}", fg=DANGER))
        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()